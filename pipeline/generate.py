"""
Stage 5: Generate.

Turn verified deals into a short, structured newsletter a business user can skim.
One LLM call writes the prose entries from the already-extracted fields; a
template fallback covers the no-key case. Writes:
  - output/newsletter.xlsx  (the deliverable, one row per deal + a prose sheet)
  - output/newsletter.md    (readable copy for the demo)
"""

from __future__ import annotations

import os
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .state import Cluster

WRITE_PROMPT = """You write a concise FMCG M&A and investment newsletter for busy
executives. For each deal below, write two sentences: what happened (who, what,
how much if known) and why it matters for the FMCG sector. Plain, factual, no
hype. Return one entry per deal in order, separated by a blank line.

Deals:
{deals}
"""


def _deal_brief(c: Cluster) -> str:
    lead = c.lead()
    flag = " [low confidence]" if c.verdict == "exhausted" else ""
    return f"- {lead.title} | type: {c.deal_type} | source: {lead.source}{flag}"


def _write_prose(deals: list[Cluster]) -> list[str]:
    if not deals:
        return []
    try:
        from .llm import get_llm

        llm = get_llm(temperature=0.3)
        block = "\n".join(_deal_brief(c) for c in deals)
        resp = llm.invoke(WRITE_PROMPT.format(deals=block))
        entries = [e.strip() for e in resp.content.split("\n\n") if e.strip()]
        if len(entries) == len(deals):
            return entries
    except Exception:
        pass
    # Template fallback.
    return [
        f"{c.lead().title}. Reported by {c.lead().source} ({c.deal_type})."
        for c in deals
    ]


def generate(deals: list[Cluster], out_dir: str = "output") -> dict:
    os.makedirs(out_dir, exist_ok=True)
    today = dt.date.today().isoformat()
    entries = _write_prose(deals)

    # Markdown copy.
    md = [f"# FMCG Deal Intelligence, {today}", ""]
    md.append(f"{len(deals)} verified deals this cycle.\n")
    for c, entry in zip(deals, entries):
        md.append(f"### {c.lead().title}")
        md.append(entry)
        md.append(f"_Source: {c.lead().source} · {c.lead().url}_")
        md.append("")
    md_text = "\n".join(md)
    md_path = os.path.join(out_dir, "newsletter.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_text)

    # Excel deliverable.
    wb = Workbook()
    ws = wb.active
    ws.title = "Newsletter"
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    cols = ["Deal", "Type", "Summary", "Source", "Confidence", "URL"]
    ws.append(cols)
    for i, _ in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = header_fill
        cell.font = header_font
    for c, entry in zip(deals, entries):
        conf = "verified" if c.verdict == "verified" else "low"
        ws.append([c.lead().title, c.deal_type, entry, c.lead().source, conf, c.lead().url])
    for col, width in zip("ABCDEF", [40, 14, 70, 20, 12, 45]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    xlsx_path = os.path.join(out_dir, "newsletter.xlsx")
    wb.save(xlsx_path)

    return {"markdown": md_text, "md_path": md_path, "xlsx_path": xlsx_path}
